#!/usr/bin/env python3
import sys
import os
import argparse
import logging

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from models import init_db, set_db_path, execute_query, execute_update
from db_init import init_and_seed, reset_db
from scheduler import ContractMonitorScheduler
from contract_scanner import (
    scan_active_contracts, get_upcoming_payment_milestones,
    get_overdue_payment_milestones, extract_milestones_for_contract
)
from reconciliation import reconcile_erp_deliveries, match_bank_payments, get_reconciliation_summary
from collection import generate_collection_reminders, send_pending_reminders, get_reminder_history
from escalation import process_escalations, notify_escalation_stakeholders, resolve_escalation, unfreeze_customer
from change_approval import (
    submit_change_request, approve_change, reject_change,
    get_pending_approvals
)
from credit_scoring import update_credit_score, batch_update_scores_from_payments, get_customer_credit_profile
from credit_downgrade import check_and_trigger_downgrade, batch_check_downgrades, get_risk_summary
from report import generate_monthly_report
from audit_log import log_operation, query_logs, batch_export_logs, get_log_statistics


def cmd_init(args):
    num_contracts = args.contracts or 500
    num_customers = args.customers or 50
    if args.reset:
        logging.info("Reset mode: clearing existing database and rebuilding...")
    init_and_seed(num_contracts, num_customers, reset=args.reset)
    logging.info(f"Database initialized with {num_contracts} contracts and {num_customers} customers")


def cmd_scan(args):
    scheduler = ContractMonitorScheduler(args.config)
    result = scheduler.run_daily_scan()
    print("\n=== Daily Scan Result ===")
    for k, v in result.items():
        print(f"  {k}: {v}")


def cmd_reconcile(args):
    scheduler = ContractMonitorScheduler(args.config)
    result = scheduler.run_bank_reconciliation()
    print("\n=== Reconciliation Result ===")
    for k, v in result.items():
        print(f"  {k}: {v}")


def cmd_collection(args):
    scheduler = ContractMonitorScheduler(args.config)
    upcoming = get_upcoming_payment_milestones(
        days_before=scheduler.config.get("collection", {}).get("pre_due_days", 7)
    )
    gen_result = generate_collection_reminders(upcoming, scheduler.config)
    send_result = send_pending_reminders(scheduler.config)
    print(f"\n=== Collection Result ===")
    print(f"  Generated: {gen_result.get('generated', 0)}")
    print(f"  Sent: {send_result.get('sent', 0)}")
    print(f"  Failed: {send_result.get('failed', 0)}")


def cmd_escalate(args):
    scheduler = ContractMonitorScheduler(args.config)
    overdue = get_overdue_payment_milestones(
        min_days=scheduler.config.get("escalation", {}).get("overdue_threshold_days", 15)
    )
    esc_result = process_escalations(overdue, scheduler.config)
    notify_result = notify_escalation_stakeholders(scheduler.config)
    print(f"\n=== Escalation Result ===")
    print(f"  Escalated: {esc_result.get('escalated', 0)}")
    print(f"  Frozen: {esc_result.get('frozen', 0)}")
    print(f"  Notified: {notify_result.get('notified', 0)}")


def cmd_change(args):
    if args.action == "submit":
        import json
        proposed = json.loads(args.proposed)
        result = submit_change_request(args.contract, args.customer, args.type, proposed, args.submitter)
        print(f"\nChange request submitted: {result}")
    elif args.action == "approve":
        result = approve_change(args.flow_id, args.approver, args.comments)
        print(f"\nChange request processed: {result}")
    elif args.action == "reject":
        result = reject_change(args.flow_id, args.approver, args.reason)
        print(f"\nChange request rejected: {result}")
    elif args.action == "list":
        pending = get_pending_approvals(role=args.role)
        print(f"\n=== Pending Approvals ({len(pending)}) ===")
        for p in pending:
            print(f"  Request: {p['request_id']}, Type: {p['change_type']}, "
                  f"Level: {p['current_approver_level']}, Contract: {p['contract_id']}")


def cmd_credit(args):
    if args.action == "update":
        result = update_credit_score(args.customer, args.event)
        print(f"\nCredit score updated: {result}")
    elif args.action == "profile":
        profile = get_customer_credit_profile(args.customer)
        if profile:
            print(f"\n=== Credit Profile: {args.customer} ===")
            c = profile["customer"]
            print(f"  Name: {c['customer_name']}")
            print(f"  Credit Level: {c['credit_level']}")
            print(f"  Credit Score: {c['credit_score']}")
            print(f"  Overdue Count: {c['overdue_count']}")
            print(f"  Order Frozen: {c['order_frozen']}")
    elif args.action == "downgrade":
        scheduler = ContractMonitorScheduler(args.config)
        results = batch_check_downgrades(scheduler.config)
        print(f"\n=== Credit Downgrade Check ===")
        print(f"  Downgraded: {len(results)}")
        for r in results:
            print(f"  Customer: {r['customer_id']}, {r['old_level']}->{r['new_level']}")


def cmd_report(args):
    scheduler = ContractMonitorScheduler(args.config)
    result = scheduler.run_monthly_report(year=args.year, month=args.month)
    print(f"\n=== Monthly Report ===")
    print(f"  Report ID: {result.get('report_id')}")
    print(f"  On-time Rate: {result['stats']['on_time_rate']}%")
    print(f"  Avg Overdue Days: {result['stats']['avg_overdue_days']}")
    print(f"  Bad Debt Rate: {result['stats']['bad_debt_rate']}%")
    print(f"  PDF: {result.get('pdf_path')}")
    print(f"  Excel: {result.get('excel_path')}")


def cmd_logs(args):
    if args.action == "query":
        logs = query_logs(
            contract_id=args.contract, customer_id=args.customer,
            operation_type=args.type, start_date=args.start, end_date=args.end,
            limit=args.limit or 100
        )
        print(f"\n=== Audit Logs ({len(logs)}) ===")
        for l in logs[:50]:
            print(f"  [{l['created_at']}] {l['operation_type']} | "
                  f"Contract: {l['contract_id']} | Customer: {l['customer_id']} | "
                  f"{l['details']}")
    elif args.action == "export":
        path = batch_export_logs(
            output_dir=args.output or "./contract_monitor/reports",
            contract_id=args.contract, customer_id=args.customer,
            start_date=args.start, end_date=args.end,
            format_type=args.format or "excel"
        )
        print(f"\nLogs exported to: {path}")
    elif args.action == "stats":
        stats = get_log_statistics(start_date=args.start, end_date=args.end)
        print(f"\n=== Log Statistics ===")
        for s in stats:
            print(f"  {s['operation_type']}: {s['cnt']}")


def cmd_full(args):
    scheduler = ContractMonitorScheduler(args.config)
    result = scheduler.run_full_daily_workflow()
    print(f"\n=== Full Daily Workflow Result ===")
    print(f"  Scan: {result['scan']}")
    print(f"  Reconciliation: {result['reconciliation']}")
    print(f"  Report: {result.get('report', {}).get('report_id') if result.get('report') else 'N/A'}")
    print(f"  Executed at: {result['executed_at']}")


def cmd_dashboard(args):
    scheduler = ContractMonitorScheduler(args.config)

    print("\n" + "=" * 60)
    print("  CONTRACT FULFILLMENT MONITORING DASHBOARD")
    print("=" * 60)

    contract_count = execute_query(
        "SELECT status, COUNT(*) as cnt FROM contracts GROUP BY status",
        fetch_all=True
    )
    print("\n[Contract Status]")
    for c in contract_count:
        print(f"  {c['status']}: {c['cnt']}")

    milestone_count = execute_query(
        "SELECT status, COUNT(*) as cnt FROM milestones GROUP BY status",
        fetch_all=True
    )
    print("\n[Milestone Status]")
    for m in milestone_count:
        print(f"  {m['status']}: {m['cnt']}")

    risk = get_risk_summary()
    print("\n[Risk Summary]")
    for r in risk.get("risk_levels", []):
        print(f"  {r['risk_level']}: {r['cnt']} customers")
    print(f"  High risk customers: {len(risk.get('high_risk_customers', []))}")

    credit_dist = risk.get("credit_distribution", [])
    print("\n[Credit Distribution]")
    for cd in credit_dist:
        print(f"  Level {cd['credit_level']}: {cd['cnt']} customers (avg score: {cd.get('avg_score', 0):.1f})")

    recon = get_reconciliation_summary()
    print("\n[Bank Matching]")
    for b in recon.get("bank_stats", []):
        status = "Matched" if b["matched"] else "Unmatched"
        print(f"  {status}: {b['cnt']} transactions, ¥{b.get('total', 0):,.2f}")

    pending_approvals = get_pending_approvals()
    print(f"\n[Pending Approvals] {len(pending_approvals)}")

    open_escalations = execute_query(
        "SELECT COUNT(*) as cnt FROM escalation_tickets WHERE status IN ('open', 'notified')",
        fetch_one=True
    )
    print(f"[Open Escalations] {open_escalations['cnt'] if open_escalations else 0}")

    print("\n" + "=" * 60)


def cmd_customers(args):
    scheduler = ContractMonitorScheduler(args.config)

    from datetime import date as date_type

    today = date_type.today()
    month_start = today.replace(day=1).isoformat()
    if today.month == 12:
        month_end = date_type(today.year + 1, 1, 1).isoformat()
    else:
        month_end = date_type(today.year, today.month + 1, 1).isoformat()

    filters = []
    params = []

    if args.credit_level:
        levels = args.credit_level.split(",")
        placeholders = ",".join(["?"] * len(levels))
        filters.append(f"cu.credit_level IN ({placeholders})")
        params.extend(levels)

    if args.manager:
        filters.append("cu.customer_id IN (SELECT customer_id FROM contracts WHERE sales_manager = ?)")
        params.append(args.manager)

    if args.frozen is not None:
        filters.append("cu.order_frozen = ?")
        params.append(1 if args.frozen else 0)

    if args.overdue_aging:
        aging = args.overdue_aging
        if aging == "0-7":
            filters.append("cu.customer_id IN (SELECT customer_id FROM milestones WHERE milestone_type='payment' AND status='overdue' AND overdue_days BETWEEN 1 AND 7)")
        elif aging == "8-15":
            filters.append("cu.customer_id IN (SELECT customer_id FROM milestones WHERE milestone_type='payment' AND status='overdue' AND overdue_days BETWEEN 8 AND 15)")
        elif aging == "16-30":
            filters.append("cu.customer_id IN (SELECT customer_id FROM milestones WHERE milestone_type='payment' AND status='overdue' AND overdue_days BETWEEN 16 AND 30)")
        elif aging == "30+":
            filters.append("cu.customer_id IN (SELECT customer_id FROM milestones WHERE milestone_type='payment' AND status='overdue' AND overdue_days > 30)")

    if args.reminder_start or args.reminder_end:
        reminder_sub = "cu.customer_id IN (SELECT customer_id FROM collection_reminders WHERE 1=1"
        reminder_params = []
        if args.reminder_start:
            reminder_sub += " AND created_at >= ?"
            reminder_params.append(args.reminder_start)
        if args.reminder_end:
            reminder_sub += " AND created_at <= ?"
            reminder_params.append(args.reminder_end)
        reminder_sub += ")"
        filters.append(reminder_sub)
        params.extend(reminder_params)

    having_parts = []
    if args.min_overdue:
        having_parts.append("overdue_amt >= ?")

    where_clause = ""
    if filters:
        where_clause = "WHERE " + " AND ".join(filters)

    having_clause = ""
    having_params = []
    if having_parts:
        having_clause = "HAVING " + " AND ".join(having_parts)
        having_params = [float(args.min_overdue)]

    sql = (
        "SELECT cu.customer_id, cu.customer_name, cu.credit_level, cu.credit_score, "
        "cu.order_frozen, cu.overdue_count, "
        "COALESCE(ms_summary.contract_count, 0) as contract_count, "
        "COALESCE(ms_summary.pending_amount, 0) as pending_amount, "
        "COALESCE(ms_summary.overdue_amt, 0) as overdue_amt, "
        "COALESCE(ms_summary.overdue_0_7, 0) as overdue_0_7, "
        "COALESCE(ms_summary.overdue_8_15, 0) as overdue_8_15, "
        "COALESCE(ms_summary.overdue_16_30, 0) as overdue_16_30, "
        "COALESCE(ms_summary.overdue_30_plus, 0) as overdue_30_plus, "
        "COALESCE(ms_summary.month_planned, 0) as month_planned, "
        "COALESCE(ms_summary.month_received, 0) as month_received, "
        "COALESCE(ms_summary.month_unreceived, 0) as month_unreceived "
        "FROM customers cu "
        "LEFT JOIN ("
        "  SELECT m.customer_id, "
        "  COUNT(DISTINCT m.contract_id) as contract_count, "
        "  SUM(CASE WHEN m.status IN ('pending','upcoming_due') THEN m.amount ELSE 0 END) as pending_amount, "
        "  SUM(CASE WHEN m.status = 'overdue' THEN m.amount ELSE 0 END) as overdue_amt, "
        "  SUM(CASE WHEN m.status = 'overdue' AND m.overdue_days BETWEEN 1 AND 7 THEN m.amount ELSE 0 END) as overdue_0_7, "
        "  SUM(CASE WHEN m.status = 'overdue' AND m.overdue_days BETWEEN 8 AND 15 THEN m.amount ELSE 0 END) as overdue_8_15, "
        "  SUM(CASE WHEN m.status = 'overdue' AND m.overdue_days BETWEEN 16 AND 30 THEN m.amount ELSE 0 END) as overdue_16_30, "
        "  SUM(CASE WHEN m.status = 'overdue' AND m.overdue_days > 30 THEN m.amount ELSE 0 END) as overdue_30_plus, "
        "  SUM(CASE WHEN m.planned_date >= ? AND m.planned_date < ? THEN m.amount ELSE 0 END) as month_planned, "
        "  SUM(CASE WHEN m.actual_date >= ? AND m.actual_date < ? THEN m.amount ELSE 0 END) as month_received, "
        "  SUM(CASE WHEN m.planned_date >= ? AND m.planned_date < ? AND m.status = 'overdue' THEN m.amount ELSE 0 END) as month_unreceived "
        "  FROM milestones m WHERE m.milestone_type = 'payment' "
        "  GROUP BY m.customer_id"
        ") ms_summary ON cu.customer_id = ms_summary.customer_id "
        f"{where_clause} "
        f"GROUP BY cu.customer_id "
        f"{having_clause} "
        "ORDER BY overdue_amt DESC"
    )
    all_params = [month_start, month_end, month_start, month_end, month_start, month_end] + params + having_params
    customer_rows = execute_query(sql, all_params, fetch_all=True)

    customer_ids = [r["customer_id"] for r in customer_rows]

    last_reminder_map = {}
    if customer_ids:
        placeholders = ",".join(["?"] * len(customer_ids))
        reminders = execute_query(
            f"SELECT customer_id, MAX(created_at) as last_reminder "
            f"FROM collection_reminders "
            f"WHERE customer_id IN ({placeholders}) "
            f"GROUP BY customer_id",
            customer_ids,
            fetch_all=True
        )
        last_reminder_map = {r["customer_id"]: r["last_reminder"] for r in reminders}

    if args.create_worklist:
        _create_worklist(customer_rows, last_reminder_map, args, scheduler.config)
        return

    print("\n" + "=" * 140)
    print("  PAYMENT RISK LEDGER (回款风险台账)")
    print("=" * 140)

    filter_parts = []
    if args.credit_level:
        filter_parts.append(f"Credit={args.credit_level}")
    if args.manager:
        filter_parts.append(f"Manager={args.manager}")
    if args.min_overdue:
        filter_parts.append(f"MinOverdue=¥{float(args.min_overdue):,.0f}")
    if args.frozen is not None:
        filter_parts.append(f"Frozen={'YES' if args.frozen else 'NO'}")
    if args.overdue_aging:
        filter_parts.append(f"Aging={args.overdue_aging}d")
    if args.reminder_start or args.reminder_end:
        rp = "Reminder:"
        if args.reminder_start:
            rp += f" from {args.reminder_start}"
        if args.reminder_end:
            rp += f" to {args.reminder_end}"
        filter_parts.append(rp)
    if filter_parts:
        print(f"  Filter: {' | '.join(filter_parts)}")

    print(f"  Total customers: {len(customer_rows)}")
    print(f"  Period: {month_start} ~ {month_end}")
    print()

    header = (f"{'ID':<10} {'Customer':<18} {'CL':<3} {'Score':<5} "
              f"{'Contracts':<5} {'Pending':>10} {'Overdue':>10} "
              f"{'0-7d':>8} {'8-15d':>8} {'16-30d':>8} {'>30d':>8} "
              f"{'MthPlan':>10} {'MthRecv':>10} {'MthUnrecv':>10} "
              f"{'Frz':<4} {'LastReminder':<20}")
    print(header)
    print("-" * 140)

    for r in customer_rows:
        frozen_str = "YES" if r["order_frozen"] else "-"
        last_rem = last_reminder_map.get(r["customer_id"], "-")
        line = (f"{r['customer_id']:<10} {r['customer_name']:<18} {r['credit_level']:<3} {r['credit_score']:<5} "
                f"{r['contract_count']:<5} {r['pending_amount']:>10,.0f} {r['overdue_amt']:>10,.0f} "
                f"{r['overdue_0_7']:>8,.0f} {r['overdue_8_15']:>8,.0f} {r['overdue_16_30']:>8,.0f} {r['overdue_30_plus']:>8,.0f} "
                f"{r['month_planned']:>10,.0f} {r['month_received']:>10,.0f} {r['month_unreceived']:>10,.0f} "
                f"{frozen_str:<4} {last_rem:<20}")
        print(line)

    totals = {
        "pending": sum(r["pending_amount"] for r in customer_rows),
        "overdue": sum(r["overdue_amt"] for r in customer_rows),
        "o07": sum(r["overdue_0_7"] for r in customer_rows),
        "o815": sum(r["overdue_8_15"] for r in customer_rows),
        "o1630": sum(r["overdue_16_30"] for r in customer_rows),
        "o30p": sum(r["overdue_30_plus"] for r in customer_rows),
        "mp": sum(r["month_planned"] for r in customer_rows),
        "mr": sum(r["month_received"] for r in customer_rows),
        "mu": sum(r["month_unreceived"] for r in customer_rows),
        "frozen": sum(1 for r in customer_rows if r["order_frozen"]),
    }

    print("-" * 140)
    print(f"{'TOTAL':<10} {'':<18} {'':<3} {'':<5} {len(customer_rows):<5} "
          f"{totals['pending']:>10,.0f} {totals['overdue']:>10,.0f} "
          f"{totals['o07']:>8,.0f} {totals['o815']:>8,.0f} {totals['o1630']:>8,.0f} {totals['o30p']:>8,.0f} "
          f"{totals['mp']:>10,.0f} {totals['mr']:>10,.0f} {totals['mu']:>10,.0f} "
          f"{totals['frozen']:<4}")
    print("=" * 140)


def cmd_customers_summary(args):
    from datetime import date as date_type

    today = date_type.today()
    month_start = today.replace(day=1).isoformat()
    if today.month == 12:
        month_end = date_type(today.year + 1, 1, 1).isoformat()
    else:
        month_end = date_type(today.year, today.month + 1, 1).isoformat()

    filters = []
    params = []

    if args.credit_level:
        levels = args.credit_level.split(",")
        placeholders = ",".join(["?"] * len(levels))
        filters.append(f"cu.credit_level IN ({placeholders})")
        params.extend(levels)

    if args.frozen is not None:
        filters.append("cu.order_frozen = ?")
        params.append(1 if args.frozen else 0)

    if args.overdue_aging:
        aging = args.overdue_aging
        if aging == "0-7":
            filters.append("cu.customer_id IN (SELECT customer_id FROM milestones WHERE milestone_type='payment' AND status='overdue' AND overdue_days BETWEEN 1 AND 7)")
        elif aging == "8-15":
            filters.append("cu.customer_id IN (SELECT customer_id FROM milestones WHERE milestone_type='payment' AND status='overdue' AND overdue_days BETWEEN 8 AND 15)")
        elif aging == "16-30":
            filters.append("cu.customer_id IN (SELECT customer_id FROM milestones WHERE milestone_type='payment' AND status='overdue' AND overdue_days BETWEEN 16 AND 30)")
        elif aging == "30+":
            filters.append("cu.customer_id IN (SELECT customer_id FROM milestones WHERE milestone_type='payment' AND status='overdue' AND overdue_days > 30)")

    where_clause = ""
    if filters:
        where_clause = "WHERE " + " AND ".join(filters)

    sql = (
        "SELECT c.sales_manager, "
        "COUNT(DISTINCT cu.customer_id) as customer_count, "
        "SUM(CASE WHEN cu.order_frozen = 1 THEN 1 ELSE 0 END) as frozen_count, "
        "COALESCE(SUM(ms_summary.overdue_amt), 0) as total_overdue, "
        "COALESCE(SUM(ms_summary.month_planned), 0) as month_planned, "
        "COALESCE(SUM(ms_summary.month_received), 0) as month_received, "
        "COALESCE(SUM(ms_summary.month_unreceived), 0) as month_unreceived, "
        "COALESCE(wl.pending_wl, 0) as pending_wl "
        "FROM customers cu "
        "JOIN contracts c ON cu.customer_id = c.customer_id AND c.status = 'active' "
        "LEFT JOIN ("
        "  SELECT m.customer_id, "
        "  SUM(CASE WHEN m.status = 'overdue' THEN m.amount ELSE 0 END) as overdue_amt, "
        "  SUM(CASE WHEN m.planned_date >= ? AND m.planned_date < ? THEN m.amount ELSE 0 END) as month_planned, "
        "  SUM(CASE WHEN m.actual_date >= ? AND m.actual_date < ? THEN m.amount ELSE 0 END) as month_received, "
        "  SUM(CASE WHEN m.planned_date >= ? AND m.planned_date < ? AND m.status = 'overdue' THEN m.amount ELSE 0 END) as month_unreceived "
        "  FROM milestones m WHERE m.milestone_type = 'payment' "
        "  GROUP BY m.customer_id"
        ") ms_summary ON cu.customer_id = ms_summary.customer_id "
        "LEFT JOIN ("
        "  SELECT assigned_manager, COUNT(*) as pending_wl FROM collection_worklists WHERE status = 'pending' GROUP BY assigned_manager"
        ") wl ON c.sales_manager = wl.assigned_manager "
        f"{where_clause} "
        "GROUP BY c.sales_manager "
        "ORDER BY total_overdue DESC"
    )
    all_params = [month_start, month_end, month_start, month_end, month_start, month_end] + params
    rows = execute_query(sql, all_params, fetch_all=True)

    print("\n" + "=" * 110)
    print("  MANAGER RISK DASHBOARD (销售经理风险看板)")
    print("=" * 110)

    filter_parts = []
    if args.credit_level:
        filter_parts.append(f"Credit={args.credit_level}")
    if args.frozen is not None:
        filter_parts.append(f"Frozen={'YES' if args.frozen else 'NO'}")
    if args.overdue_aging:
        filter_parts.append(f"Aging={args.overdue_aging}d")
    if filter_parts:
        print(f"  Filter: {' | '.join(filter_parts)}")

    print(f"  Period: {month_start} ~ {month_end}")
    print()

    header = (f"{'Manager':<12} {'Custs':>5} {'Frozen':>6} {'TotalOverdue':>14} "
              f"{'MthPlan':>12} {'MthRecv':>12} {'MthUnrecv':>12} {'AvgRate%':>8} {'PendingWL':>9}")
    print(header)
    print("-" * 110)

    total_custs = 0
    total_frozen = 0
    total_overdue = 0
    total_month_planned = 0
    total_month_received = 0
    total_month_unreceived = 0
    total_pending_wl = 0

    for r in rows:
        avg_rate = (r["month_received"] / r["month_planned"] * 100) if r["month_planned"] > 0 else 0.0
        print(f"{r['sales_manager']:<12} {r['customer_count']:>5} {r['frozen_count']:>6} "
              f"{r['total_overdue']:>14,.0f} {r['month_planned']:>12,.0f} "
              f"{r['month_received']:>12,.0f} {r['month_unreceived']:>12,.0f} "
              f"{avg_rate:>8.1f} {r['pending_wl']:>9}")
        total_custs += r["customer_count"]
        total_frozen += r["frozen_count"]
        total_overdue += r["total_overdue"]
        total_month_planned += r["month_planned"]
        total_month_received += r["month_received"]
        total_month_unreceived += r["month_unreceived"]
        total_pending_wl += r["pending_wl"]

    total_avg_rate = (total_month_received / total_month_planned * 100) if total_month_planned > 0 else 0.0

    print("-" * 110)
    print(f"{'TOTAL':<12} {total_custs:>5} {total_frozen:>6} "
          f"{total_overdue:>14,.0f} {total_month_planned:>12,.0f} "
          f"{total_month_received:>12,.0f} {total_month_unreceived:>12,.0f} "
          f"{total_avg_rate:>8.1f} {total_pending_wl:>9}")
    print("=" * 110)


def _create_worklist(customer_rows, last_reminder_map, args, config):
    from models import execute_many
    import uuid
    from datetime import datetime, timedelta, date as date_type

    items = []
    for r in customer_rows:
        if r["overdue_amt"] <= 0:
            continue
        cl = r["credit_level"]
        if cl in ("A", "B"):
            tactic = "友好提醒：客户信用良好，建议温和沟通确认付款计划"
        elif cl == "C":
            tactic = "正式催收：发送书面催款函，要求给出明确付款日期"
        else:
            tactic = "强硬措施：发送律师函预警，暂停新订单审批，准备法务介入"

        manager_rows = execute_query(
            "SELECT sales_manager FROM contracts WHERE customer_id = ? AND status = 'active' LIMIT 1",
            [r["customer_id"]],
            fetch_all=True
        )
        manager = manager_rows[0]["sales_manager"] if manager_rows else ""

        overdue_items = execute_query(
            "SELECT contract_no, amount, overdue_days, description FROM milestones "
            "WHERE customer_id = ? AND milestone_type = 'payment' AND status = 'overdue' "
            "ORDER BY overdue_days DESC",
            [r["customer_id"]],
            fetch_all=True
        )

        contract_info = "; ".join(
            f"{m['contract_no']}({m['description']}:{m['amount']:,.0f}元/逾期{m['overdue_days']}天)"
            for m in overdue_items[:3]
        )

        deadline = (date_type.today() + timedelta(days=7)).isoformat()
        item_id = f"WL{uuid.uuid4().hex[:12]}"

        items.append((
            item_id, r["customer_id"], r["customer_name"], r["credit_level"],
            r["overdue_amt"], contract_info, tactic, manager,
            deadline, "pending", datetime.now().isoformat()
        ))

    if not items:
        print("\nNo overdue customers found for worklist creation")
        return

    execute_many(
        "INSERT OR IGNORE INTO collection_worklists "
        "(item_id, customer_id, customer_name, credit_level, overdue_amount, "
        "contract_details, suggested_tactic, assigned_manager, "
        "deadline, status, created_at) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        items
    )

    print(f"\n=== Worklist Created ===")
    print(f"  Items: {len(items)}")
    print(f"  Deadline: {(date_type.today() + timedelta(days=7)).isoformat()}")
    print()
    for it in items:
        print(f"  [{it[3]}] {it[2]} | ¥{it[4]:,.0f} | {it[7]} | {it[6][:30]}...")
    print(f"\nUse 'python main.py worklist' to view and manage worklist items")


def cmd_worklist(args):
    if args.action == "list":
        sql = "SELECT * FROM collection_worklists WHERE 1=1"
        params = []
        if args.manager:
            sql += " AND assigned_manager = ?"
            params.append(args.manager)
        if args.status:
            sql += " AND status = ?"
            params.append(args.status)
        if args.deadline_start:
            sql += " AND deadline >= ?"
            params.append(args.deadline_start)
        if args.deadline_end:
            sql += " AND deadline <= ?"
            params.append(args.deadline_end)
        if args.min_overdue:
            sql += " AND overdue_amount >= ?"
            params.append(float(args.min_overdue))
        if args.credit_level:
            levels = args.credit_level.split(",")
            placeholders = ",".join(["?"] * len(levels))
            sql += f" AND credit_level IN ({placeholders})"
            params.extend(levels)
        sql += " ORDER BY created_at DESC LIMIT 200"

        items = execute_query(sql, params, fetch_all=True)
        print(f"\n=== Collection Worklist ({len(items)} items) ===")
        if not items:
            print("  No items found")
            return

        print(f"{'ItemID':<16} {'Customer':<18} {'CL':<3} {'Overdue':>10} "
              f"{'Manager':<10} {'Deadline':<12} {'Status':<8} {'Tactic':<30}")
        print("-" * 120)
        for it in items:
            print(f"{it['item_id']:<16} {it['customer_name']:<18} {it['credit_level']:<3} "
                  f"{it['overdue_amount']:>10,.0f} {it['assigned_manager']:<10} "
                  f"{it['deadline']:<12} {it['status']:<8} {it['suggested_tactic'][:28]:<30}")
        print("-" * 120)

        if args.detail:
            print("\n--- Overdue Contract Details ---")
            for it in items:
                if it["contract_details"]:
                    print(f"  [{it['item_id']}] {it['customer_name']}: {it['contract_details']}")
                else:
                    overdue_rows = execute_query(
                        "SELECT contract_no, amount, overdue_days, description FROM milestones "
                        "WHERE customer_id = ? AND milestone_type = 'payment' AND status = 'overdue' "
                        "ORDER BY overdue_days DESC",
                        [it["customer_id"]],
                        fetch_all=True
                    )
                    details = "; ".join(
                        f"{m['contract_no']}({m['description']}:{m['amount']:,.0f}元/逾期{m['overdue_days']}天)"
                        for m in overdue_rows
                    )
                    print(f"  [{it['item_id']}] {it['customer_name']}: {details}")

    elif args.action == "export":
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        sql = "SELECT * FROM collection_worklists WHERE 1=1"
        params = []
        if args.status:
            sql += " AND status = ?"
            params.append(args.status)
        if args.manager:
            sql += " AND assigned_manager = ?"
            params.append(args.manager)
        sql += " ORDER BY status ASC, created_at DESC"

        items = execute_query(sql, params, fetch_all=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Collection Worklist"

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        headers = ["Customer", "Credit Level", "Contract Details", "Overdue Amount",
                    "Suggested Tactic", "Assigned Manager", "Deadline",
                    "Status", "Result", "Note"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for it in items:
            ws.append([
                it["customer_name"], it["credit_level"], it["contract_details"],
                it["overdue_amount"], it["suggested_tactic"], it["assigned_manager"],
                it["deadline"], it["status"],
                "已处理" if it["status"] == "done" else ("处理失败" if it["status"] == "failed" else ""),
                it.get("note", "")
            ])
            for cell in ws[ws.max_row]:
                cell.border = thin_border

        widths = [18, 10, 40, 14, 35, 14, 12, 8, 8, 20]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width

        from datetime import date as date_type
        filename = f"催收清单_{date_type.today().isoformat()}.xlsx"
        filepath = os.path.join("./contract_monitor/reports", filename)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        wb.save(filepath)
        print(f"Worklist exported to: {filepath} ({len(items)} items)")

    elif args.action in ("mark-done", "mark-fail"):
        if not args.item_id:
            print("Error: --item-id is required for status update")
            return
        new_status = "done" if args.action == "mark-done" else "failed"
        from datetime import datetime
        execute_update(
            "UPDATE collection_worklists SET status = ?, note = ?, processed_at = ? WHERE item_id = ?",
            [new_status, args.note, datetime.now().isoformat(), args.item_id]
        )
        print(f"Worklist item {args.item_id} marked as {new_status}")

        if new_status == "done":
            item = execute_query(
                "SELECT customer_id FROM collection_worklists WHERE item_id = ?",
                [args.item_id], fetch_one=True
            )
            if item:
                from audit_log import log_operation
                log_operation("worklist_completed", "", item["customer_id"],
                              f"催收清单处理完成: {args.item_id}")


def main():
    parser = argparse.ArgumentParser(
        description="Contract Fulfillment Monitoring & Intelligent Collection System"
    )
    parser.add_argument("--config", default="config.yaml", help="Config file path")

    subparsers = parser.add_subparsers(dest="command", help="Available commands")

    p_init = subparsers.add_parser("init", help="Initialize database with sample data")
    p_init.add_argument("--contracts", type=int, default=500, help="Number of sample contracts")
    p_init.add_argument("--customers", type=int, default=50, help="Number of sample customers")
    p_init.add_argument("--reset", action="store_true", help="Reset database before seeding (clean rebuild)")

    p_scan = subparsers.add_parser("scan", help="Run daily contract scan")
    p_reconcile = subparsers.add_parser("reconcile", help="Run bank reconciliation")
    p_collection = subparsers.add_parser("collection", help="Process collection reminders")
    p_escalate = subparsers.add_parser("escalate", help="Process overdue escalations")

    p_change = subparsers.add_parser("change", help="Manage change requests")
    p_change.add_argument("action", choices=["submit", "approve", "reject", "list"])
    p_change.add_argument("--contract", help="Contract ID")
    p_change.add_argument("--customer", help="Customer ID")
    p_change.add_argument("--type", choices=["extension", "installment"], help="Change type")
    p_change.add_argument("--proposed", help="Proposed data (JSON)")
    p_change.add_argument("--submitter", default="system")
    p_change.add_argument("--flow-id", help="Approval flow ID")
    p_change.add_argument("--approver", help="Approver ID")
    p_change.add_argument("--comments", default="")
    p_change.add_argument("--reason", default="")
    p_change.add_argument("--role", help="Filter by approver role")

    p_credit = subparsers.add_parser("credit", help="Manage credit scores")
    p_credit.add_argument("action", choices=["update", "profile", "downgrade"])
    p_credit.add_argument("--customer", help="Customer ID")
    p_credit.add_argument("--event", choices=["on_time", "minor_late", "major_late"])

    p_report = subparsers.add_parser("report", help="Generate monthly report")
    p_report.add_argument("--year", type=int)
    p_report.add_argument("--month", type=int)

    p_logs = subparsers.add_parser("logs", help="Query and export audit logs")
    p_logs.add_argument("action", choices=["query", "export", "stats"])
    p_logs.add_argument("--contract", help="Filter by contract ID")
    p_logs.add_argument("--customer", help="Filter by customer ID")
    p_logs.add_argument("--type", help="Filter by operation type")
    p_logs.add_argument("--start", help="Start date")
    p_logs.add_argument("--end", help="End date")
    p_logs.add_argument("--limit", type=int, default=100)
    p_logs.add_argument("--output", help="Export output directory")
    p_logs.add_argument("--format", choices=["excel", "csv"], default="excel")

    p_full = subparsers.add_parser("full", help="Run full daily workflow")
    p_dashboard = subparsers.add_parser("dashboard", help="Show monitoring dashboard")

    p_customers = subparsers.add_parser("customers", help="Payment risk ledger with filters")
    p_customers.add_argument("--credit-level", help="Filter by credit level(s), e.g. A,B or C,D")
    p_customers.add_argument("--manager", help="Filter by sales manager name")
    p_customers.add_argument("--min-overdue", help="Filter by minimum overdue amount")
    p_customers.add_argument("--frozen", type=int, default=None, help="Filter frozen status: 1=frozen, 0=not frozen")
    p_customers.add_argument("--overdue-aging", choices=["0-7", "8-15", "16-30", "30+"], help="Filter by overdue aging bucket")
    p_customers.add_argument("--reminder-start", help="Filter by last reminder start date (YYYY-MM-DD)")
    p_customers.add_argument("--reminder-end", help="Filter by last reminder end date (YYYY-MM-DD)")
    p_customers.add_argument("--create-worklist", action="store_true", help="Generate sales follow-up worklist for filtered customers")

    p_customers_summary = subparsers.add_parser("customers-summary", help="Sales manager risk dashboard")
    p_customers_summary.add_argument("--credit-level", help="Filter by credit level(s), e.g. A,B or C,D")
    p_customers_summary.add_argument("--overdue-aging", choices=["0-7", "8-15", "16-30", "30+"], help="Filter by overdue aging bucket")
    p_customers_summary.add_argument("--frozen", type=int, default=None, help="Filter frozen status: 1=frozen, 0=not frozen")

    p_worklist = subparsers.add_parser("worklist", help="View and manage collection worklists")
    p_worklist.add_argument("action", choices=["list", "mark-done", "mark-fail", "export"], help="Worklist action")
    p_worklist.add_argument("--item-id", help="Worklist item ID to update")
    p_worklist.add_argument("--manager", help="Filter by assigned manager")
    p_worklist.add_argument("--status", choices=["pending", "done", "failed"], help="Filter by status")
    p_worklist.add_argument("--note", default="", help="Note for status update")
    p_worklist.add_argument("--detail", action="store_true", help="Show overdue contract details in list")
    p_worklist.add_argument("--deadline-start", help="Filter deadline start date (YYYY-MM-DD)")
    p_worklist.add_argument("--deadline-end", help="Filter deadline end date (YYYY-MM-DD)")
    p_worklist.add_argument("--min-overdue", help="Filter by minimum overdue amount")
    p_worklist.add_argument("--credit-level", help="Filter by credit level(s), e.g. A,B or C,D")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return

    commands = {
        "init": cmd_init,
        "scan": cmd_scan,
        "reconcile": cmd_reconcile,
        "collection": cmd_collection,
        "escalate": cmd_escalate,
        "change": cmd_change,
        "credit": cmd_credit,
        "report": cmd_report,
        "logs": cmd_logs,
        "full": cmd_full,
        "dashboard": cmd_dashboard,
        "customers": cmd_customers,
        "customers-summary": cmd_customers_summary,
        "worklist": cmd_worklist,
    }

    if args.command in commands:
        commands[args.command](args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
