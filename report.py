import os
import json
import logging
from datetime import datetime, date, timedelta
from models import execute_query, execute_update

logger = logging.getLogger(__name__)

REPORT_DIR = "./contract_monitor/reports"


def _ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def generate_monthly_report(year=None, month=None, config=None):
    if config is None:
        config = {}

    if year is None or month is None:
        today = date.today()
        if today.day == 1:
            last_month = today.replace(day=1) - timedelta(days=1)
            year = last_month.year
            month = last_month.month
        else:
            year = today.year
            month = today.month

    output_dir = config.get("report", {}).get("output_dir", REPORT_DIR)
    _ensure_dir(output_dir)

    start_date = date(year, month, 1).isoformat()
    if month == 12:
        end_date = date(year + 1, 1, 1).isoformat()
    else:
        end_date = date(year, month + 1, 1).isoformat()

    logger.info(f"Generating monthly report for {year}-{month:02d}")

    stats = _compute_report_stats(start_date, end_date)

    pdf_path = None
    excel_path = None

    if config.get("report", {}).get("pdf_enabled", True):
        try:
            pdf_path = _generate_pdf_report(stats, year, month, output_dir)
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            pdf_path = None

    if config.get("report", {}).get("excel_enabled", True):
        try:
            excel_path = _generate_excel_report(stats, year, month, output_dir)
        except Exception as e:
            logger.error(f"Excel generation failed: {e}")
            excel_path = None

    report_id = f"RPT{year}{month:02d}"
    execute_update(
        "INSERT OR REPLACE INTO monthly_reports "
        "(report_id, report_month, on_time_rate, avg_overdue_days, bad_debt_rate, "
        "credit_distribution, generated_date, pdf_path, excel_path) "
        "VALUES (?,?,?,?,?,?,?,?,?)",
        [report_id, f"{year}-{month:02d}", stats["on_time_rate"],
         stats["avg_overdue_days"], stats["bad_debt_rate"],
         json.dumps(stats["credit_distribution"], ensure_ascii=False),
         datetime.now().isoformat(), pdf_path, excel_path]
    )

    _log_operation("report_generated", "", "",
                   f"月度报告生成: {year}-{month:02d}, PDF: {pdf_path}, Excel: {excel_path}")

    logger.info(f"Monthly report generated: PDF={pdf_path}, Excel={excel_path}")
    return {
        "report_id": report_id,
        "year": year, "month": month,
        "stats": stats,
        "pdf_path": pdf_path,
        "excel_path": excel_path
    }


def _compute_report_stats(start_date, end_date):
    start_d = date.fromisoformat(start_date)
    end_d = date.fromisoformat(end_date)

    total_payment_milestones = execute_query(
        "SELECT COUNT(*) as cnt FROM milestones "
        "WHERE milestone_type = 'payment' "
        "AND planned_date >= ? AND planned_date < ?",
        [start_date, end_date],
        fetch_one=True
    )

    completed_on_time = execute_query(
        "SELECT COUNT(*) as cnt FROM milestones "
        "WHERE milestone_type = 'payment' AND status = 'completed' "
        "AND overdue_days = 0 "
        "AND actual_date >= ? AND actual_date < ?",
        [start_date, end_date],
        fetch_one=True
    )

    completed_late = execute_query(
        "SELECT COUNT(*) as cnt, AVG(overdue_days) as avg_days "
        "FROM milestones "
        "WHERE milestone_type = 'payment' AND status = 'completed' "
        "AND overdue_days > 0 "
        "AND actual_date >= ? AND actual_date < ?",
        [start_date, end_date],
        fetch_one=True
    )

    overdue = execute_query(
        "SELECT COUNT(*) as cnt, COALESCE(SUM(amount), 0) as total "
        "FROM milestones "
        "WHERE milestone_type = 'payment' AND status = 'overdue' "
        "AND planned_date >= ? AND planned_date < ?",
        [start_date, end_date],
        fetch_one=True
    )

    total_completed = (completed_on_time["cnt"] or 0) + (completed_late["cnt"] or 0)
    on_time_rate = (completed_on_time["cnt"] or 0) / max(total_completed, 1) * 100
    avg_overdue_days = completed_late["avg_days"] or 0

    total_planned_amount = execute_query(
        "SELECT COALESCE(SUM(amount), 0) as total FROM milestones "
        "WHERE milestone_type = 'payment' "
        "AND planned_date >= ? AND planned_date < ?",
        [start_date, end_date],
        fetch_one=True
    )

    bad_debt_rate = (overdue["total"] or 0) / max(total_planned_amount["total"] or 1, 1) * 100

    credit_dist = execute_query(
        "SELECT cu.credit_level, COUNT(DISTINCT cu.customer_id) as cnt "
        "FROM customers cu "
        "JOIN milestones m ON cu.customer_id = m.customer_id "
        "WHERE m.milestone_type = 'payment' "
        "AND (m.planned_date >= ? AND m.planned_date < ? "
        "     OR (m.actual_date >= ? AND m.actual_date < ?)) "
        "GROUP BY cu.credit_level ORDER BY cu.credit_level",
        [start_date, end_date, start_date, end_date],
        fetch_all=True
    )

    credit_distribution = {item["credit_level"]: item["cnt"] for item in credit_dist}

    top_overdue_customers = execute_query(
        "SELECT cu.customer_id, cu.customer_name, cu.credit_level, "
        "COUNT(m.milestone_id) as overdue_cnt, SUM(m.amount) as overdue_total "
        "FROM customers cu "
        "JOIN milestones m ON cu.customer_id = m.customer_id "
        "WHERE m.status = 'overdue' AND m.milestone_type = 'payment' "
        "AND m.planned_date >= ? AND m.planned_date < ? "
        "GROUP BY cu.customer_id "
        "ORDER BY overdue_total DESC LIMIT 10",
        [start_date, end_date],
        fetch_all=True
    )

    monthly_trend = []
    for i in range(6):
        offset_months = i + 1
        target = _month_offset(start_d, -offset_months)
        m_start = target.isoformat()
        m_end_month = _month_offset(target, 1)
        m_end = m_end_month.isoformat()

        m_completed = execute_query(
            "SELECT COUNT(*) as cnt FROM milestones "
            "WHERE milestone_type = 'payment' AND status = 'completed' "
            "AND actual_date >= ? AND actual_date < ?",
            [m_start, m_end],
            fetch_one=True
        )
        m_overdue = execute_query(
            "SELECT COUNT(*) as cnt FROM milestones "
            "WHERE milestone_type = 'payment' AND status = 'overdue' "
            "AND planned_date >= ? AND planned_date < ?",
            [m_start, m_end],
            fetch_one=True
        )
        monthly_trend.append({
            "month": f"{target.year}-{target.month:02d}",
            "completed": m_completed["cnt"] or 0,
            "overdue": m_overdue["cnt"] or 0
        })

    monthly_trend.reverse()

    customer_payment_analysis = execute_query(
        "SELECT cu.customer_id, cu.customer_name, cu.credit_level, "
        "COALESCE(ms_summary.sales_manager, '') as sales_manager, "
        "COALESCE(ms_summary.month_planned, 0) as month_planned, "
        "COALESCE(ms_summary.month_received, 0) as month_received, "
        "COALESCE(ms_summary.month_overdue, 0) as month_overdue, "
        "CASE WHEN COALESCE(ms_summary.month_planned, 0) > 0 "
        "  THEN ROUND(COALESCE(ms_summary.month_received, 0) * 100.0 "
        "    / ms_summary.month_planned, 2) "
        "  ELSE 0 END as payment_rate "
        "FROM customers cu "
        "LEFT JOIN ("
        "  SELECT m.customer_id, "
        "  (SELECT c2.sales_manager FROM contracts c2 WHERE c2.customer_id = m.customer_id AND c2.status = 'active' LIMIT 1) as sales_manager, "
        "  SUM(CASE WHEN m.planned_date >= ? AND m.planned_date < ? THEN m.amount ELSE 0 END) as month_planned, "
        "  SUM(CASE WHEN m.actual_date >= ? AND m.actual_date < ? AND m.status = 'completed' THEN m.amount ELSE 0 END) as month_received, "
        "  SUM(CASE WHEN m.planned_date >= ? AND m.planned_date < ? AND m.status = 'overdue' THEN m.amount ELSE 0 END) as month_overdue "
        "  FROM milestones m WHERE m.milestone_type = 'payment' "
        "  AND (m.planned_date >= ? AND m.planned_date < ? "
        "       OR (m.actual_date >= ? AND m.actual_date < ?)) "
        "  GROUP BY m.customer_id"
        ") ms_summary ON cu.customer_id = ms_summary.customer_id "
        "WHERE ms_summary.customer_id IS NOT NULL "
        "ORDER BY month_overdue DESC",
        [start_date, end_date, start_date, end_date, start_date, end_date,
         start_date, end_date, start_date, end_date],
        fetch_all=True
    )

    credit_changes = execute_query(
        "SELECT ch.customer_id, ch.old_level, ch.new_level, ch.reason "
        "FROM credit_history ch "
        "WHERE ch.change_date >= ? AND ch.change_date < ? "
        "ORDER BY ch.change_date DESC",
        [start_date, end_date],
        fetch_all=True
    )
    credit_change_map = {}
    for ch in credit_changes:
        cid = ch["customer_id"]
        if cid not in credit_change_map:
            credit_change_map[cid] = []
        credit_change_map[cid].append(f"{ch['old_level']}->{ch['new_level']}")

    for cp in customer_payment_analysis:
        cp["credit_change"] = "; ".join(credit_change_map.get(cp["customer_id"], ["-"]))

    planned_details = execute_query(
        "SELECT m.contract_no, cu.customer_name, m.description, "
        "m.planned_date, m.actual_date, m.amount, m.overdue_days, m.status "
        "FROM milestones m "
        "JOIN customers cu ON m.customer_id = cu.customer_id "
        "WHERE m.milestone_type = 'payment' "
        "AND m.planned_date >= ? AND m.planned_date < ? "
        "ORDER BY m.planned_date ASC",
        [start_date, end_date],
        fetch_all=True
    )

    actual_details = execute_query(
        "SELECT m.contract_no, cu.customer_name, m.description, "
        "m.planned_date, m.actual_date, m.amount, m.overdue_days, m.status "
        "FROM milestones m "
        "JOIN customers cu ON m.customer_id = cu.customer_id "
        "WHERE m.milestone_type = 'payment' AND m.actual_date IS NOT NULL "
        "AND m.actual_date >= ? AND m.actual_date < ? "
        "ORDER BY m.actual_date ASC",
        [start_date, end_date],
        fetch_all=True
    )

    overdue_details = execute_query(
        "SELECT m.contract_no, cu.customer_name, m.description, "
        "m.planned_date, m.actual_date, m.amount, m.overdue_days, m.status "
        "FROM milestones m "
        "JOIN customers cu ON m.customer_id = cu.customer_id "
        "WHERE m.milestone_type = 'payment' AND m.status = 'overdue' "
        "AND m.planned_date >= ? AND m.planned_date < ? "
        "ORDER BY m.overdue_days DESC",
        [start_date, end_date],
        fetch_all=True
    )

    manager_analysis = execute_query(
        "SELECT c.sales_manager, "
        "COUNT(DISTINCT cu.customer_id) as customer_count, "
        "SUM(CASE WHEN cu.order_frozen = 1 THEN 1 ELSE 0 END) as frozen_count, "
        "COALESCE(SUM(ms_summary.month_planned), 0) as month_planned, "
        "COALESCE(SUM(ms_summary.month_received), 0) as month_received, "
        "COALESCE(SUM(ms_summary.month_overdue), 0) as month_overdue, "
        "CASE WHEN COALESCE(SUM(ms_summary.month_planned), 0) > 0 "
        "  THEN ROUND(COALESCE(SUM(ms_summary.month_received), 0) * 100.0 "
        "    / SUM(ms_summary.month_planned), 2) "
        "  ELSE 0 END as payment_rate, "
        "COALESCE(wl_done.done_count, 0) as wl_done_count, "
        "COALESCE(wl_total.total_count, 0) as wl_total_count, "
        "CASE WHEN COALESCE(wl_total.total_count, 0) > 0 "
        "  THEN ROUND(COALESCE(wl_done.done_count, 0) * 100.0 / wl_total.total_count, 2) "
        "  ELSE 0 END as collection_rate "
        "FROM customers cu "
        "JOIN contracts c ON cu.customer_id = c.customer_id AND c.status = 'active' "
        "LEFT JOIN ("
        "  SELECT m.customer_id, "
        "  SUM(CASE WHEN m.planned_date >= ? AND m.planned_date < ? THEN m.amount ELSE 0 END) as month_planned, "
        "  SUM(CASE WHEN m.actual_date >= ? AND m.actual_date < ? AND m.status = 'completed' THEN m.amount ELSE 0 END) as month_received, "
        "  SUM(CASE WHEN m.planned_date >= ? AND m.planned_date < ? AND m.status = 'overdue' THEN m.amount ELSE 0 END) as month_overdue "
        "  FROM milestones m WHERE m.milestone_type = 'payment' "
        "  AND (m.planned_date >= ? AND m.planned_date < ? "
        "       OR (m.actual_date >= ? AND m.actual_date < ?)) "
        "  GROUP BY m.customer_id"
        ") ms_summary ON cu.customer_id = ms_summary.customer_id "
        "LEFT JOIN ("
        "  SELECT assigned_manager, COUNT(*) as done_count FROM collection_worklists WHERE status = 'done' GROUP BY assigned_manager"
        ") wl_done ON c.sales_manager = wl_done.assigned_manager "
        "LEFT JOIN ("
        "  SELECT assigned_manager, COUNT(*) as total_count FROM collection_worklists GROUP BY assigned_manager"
        ") wl_total ON c.sales_manager = wl_total.assigned_manager "
        "WHERE ms_summary.customer_id IS NOT NULL "
        "GROUP BY c.sales_manager "
        "ORDER BY month_overdue DESC",
        [start_date, end_date, start_date, end_date, start_date, end_date,
         start_date, end_date, start_date, end_date],
        fetch_all=True
    )

    return {
        "on_time_rate": round(on_time_rate, 2),
        "avg_overdue_days": round(avg_overdue_days, 1),
        "bad_debt_rate": round(bad_debt_rate, 2),
        "credit_distribution": credit_distribution,
        "total_payment_milestones": total_payment_milestones["cnt"] or 0,
        "completed_on_time": completed_on_time["cnt"] or 0,
        "completed_late": completed_late["cnt"] or 0,
        "overdue_count": overdue["cnt"] or 0,
        "overdue_amount": overdue["total"] or 0,
        "top_overdue_customers": top_overdue_customers,
        "monthly_trend": monthly_trend,
        "customer_payment_analysis": customer_payment_analysis,
        "manager_payment_analysis": manager_analysis,
        "planned_details": planned_details,
        "actual_details": actual_details,
        "overdue_details": overdue_details,
    }


def _month_offset(base_date, months):
    month = base_date.month - 1 + months
    year = base_date.year + month // 12
    month = month % 12 + 1
    day = min(base_date.day, 28)
    return date(year, month, day)


def _generate_pdf_report(stats, year, month, output_dir):
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib import font_manager
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError as e:
        logger.error(f"Missing dependency for PDF generation: {e}")
        return None

    _setup_chinese_fonts()

    pdf_filename = f"履约分析报告_{year}年{month:02d}月.pdf"
    pdf_path = os.path.join(output_dir, pdf_filename)

    doc = SimpleDocTemplate(pdf_path, pagesize=A4,
                            topMargin=20 * mm, bottomMargin=20 * mm,
                            leftMargin=15 * mm, rightMargin=15 * mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ChineseTitle", parent=styles["Title"], fontSize=18)
    heading_style = ParagraphStyle("ChineseHeading", parent=styles["Heading2"], fontSize=14)
    body_style = ParagraphStyle("ChineseBody", parent=styles["Normal"], fontSize=10)

    elements = []

    elements.append(Paragraph(f"合同履约分析报告 - {year}年{month:02d}月", title_style))
    elements.append(Spacer(1, 10 * mm))

    summary_data = [
        ["指标", "数值"],
        ["按时付款率", f"{stats['on_time_rate']}%"],
        ["平均逾期天数", f"{stats['avg_overdue_days']}天"],
        ["坏账率", f"{stats['bad_debt_rate']}%"],
        ["总付款里程碑数", str(stats['total_payment_milestones'])],
        ["按时完成数", str(stats['completed_on_time'])],
        ["逾期完成数", str(stats['completed_late'])],
        ["当前逾期数", str(stats['overdue_count'])],
        ["逾期总金额", f"¥{stats['overdue_amount']:,.2f}"],
    ]

    table = Table(summary_data, colWidths=[60 * mm, 80 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#ECF0F1")]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 10 * mm))

    try:
        chart_paths = _generate_charts(stats, year, month, output_dir)
        for chart_path in chart_paths:
            if os.path.exists(chart_path):
                elements.append(Paragraph(chart_path.split("/")[-1].replace(".png", ""), heading_style))
                elements.append(Image(chart_path, width=160 * mm, height=90 * mm))
                elements.append(Spacer(1, 5 * mm))
    except Exception as e:
        logger.error(f"Chart generation failed: {e}")

    if stats.get("top_overdue_customers"):
        elements.append(Paragraph("Top 10逾期客户", heading_style))
        top_data = [["客户名称", "信用等级", "逾期次数", "逾期金额"]]
        for c in stats["top_overdue_customers"]:
            top_data.append([
                c["customer_name"], c["credit_level"],
                str(c["overdue_cnt"]), f"¥{c['overdue_total']:,.2f}"
            ])
        top_table = Table(top_data, colWidths=[50 * mm, 30 * mm, 30 * mm, 40 * mm])
        top_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E74C3C")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(top_table)

    doc.build(elements)
    logger.info(f"PDF report generated: {pdf_path}")
    return pdf_path


def _setup_chinese_fonts():
    pass


def _generate_charts(stats, year, month, output_dir):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    chart_paths = []

    plt.rcParams["font.sans-serif"] = ["Arial Unicode MS", "SimHei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    credit_dist = stats.get("credit_distribution", {})
    if credit_dist:
        fig, ax = plt.subplots(figsize=(8, 4))
        levels = list(credit_dist.keys())
        counts = list(credit_dist.values())
        bar_colors = ["#2ECC71", "#3498DB", "#F39C12", "#E74C3C"]
        ax.bar(levels, counts, color=bar_colors[:len(levels)])
        ax.set_xlabel("Credit Level")
        ax.set_ylabel("Customer Count")
        ax.set_title("Credit Level Distribution")
        for i, v in enumerate(counts):
            ax.text(i, v + 0.5, str(v), ha="center", va="bottom")
        chart_path = os.path.join(output_dir, f"credit_distribution_{year}{month:02d}.png")
        fig.savefig(chart_path, dpi=100, bbox_inches="tight")
        plt.close(fig)
        chart_paths.append(chart_path)

    trend = stats.get("monthly_trend", [])
    if trend:
        fig, ax = plt.subplots(figsize=(8, 4))
        months_list = [t["month"] for t in trend]
        completed_list = [t["completed"] for t in trend]
        overdue_list = [t["overdue"] for t in trend]
        x = range(len(months_list))
        width = 0.35
        ax.bar([i - width / 2 for i in x], completed_list, width, label="Completed", color="#2ECC71")
        ax.bar([i + width / 2 for i in x], overdue_list, width, label="Overdue", color="#E74C3C")
        ax.set_xlabel("Month")
        ax.set_ylabel("Count")
        ax.set_title("Monthly Payment Trend (Last 6 Months)")
        ax.set_xticks(list(x))
        ax.set_xticklabels(months_list, rotation=45)
        ax.legend()
        chart_path = os.path.join(output_dir, f"monthly_trend_{year}{month:02d}.png")
        fig.savefig(chart_path, dpi=100, bbox_inches="tight")
        plt.close(fig)
        chart_paths.append(chart_path)

    return chart_paths


def _generate_excel_report(stats, year, month, output_dir):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import BarChart, PieChart, Reference
    except ImportError as e:
        logger.error(f"Missing openpyxl: {e}")
        return None

    excel_filename = f"履约分析报告_{year}年{month:02d}月.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)

    wb = Workbook()

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Contract Fulfillment Analysis Report", f"{year}-{month:02d}"])
    ws_summary.merge_cells("A1:D1")
    ws_summary.append([])

    summary_headers = ["Metric", "Value"]
    ws_summary.append(summary_headers)
    for cell in ws_summary[3]:
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    metrics = [
        ("On-time Payment Rate", f"{stats['on_time_rate']}%"),
        ("Average Overdue Days", f"{stats['avg_overdue_days']} days"),
        ("Bad Debt Rate", f"{stats['bad_debt_rate']}%"),
        ("Total Payment Milestones", stats["total_payment_milestones"]),
        ("Completed On-time", stats["completed_on_time"]),
        ("Completed Late", stats["completed_late"]),
        ("Currently Overdue", stats["overdue_count"]),
        ("Overdue Amount", f"¥{stats['overdue_amount']:,.2f}"),
    ]
    for metric in metrics:
        ws_summary.append(list(metric))
        for cell in ws_summary[ws_summary.max_row]:
            cell.border = thin_border

    ws_credit = wb.create_sheet("Credit Distribution")
    ws_credit.append(["Credit Level", "Customer Count"])
    for cell in ws_credit[1]:
        cell.font = header_font
        cell.fill = header_fill

    credit_dist = stats.get("credit_distribution", {})
    for level, count in sorted(credit_dist.items()):
        ws_credit.append([level, count])

    if credit_dist:
        pie = PieChart()
        pie.title = "Credit Level Distribution"
        data = Reference(ws_credit, min_col=2, min_row=1, max_row=ws_credit.max_row)
        labels = Reference(ws_credit, min_col=1, min_row=2, max_row=ws_credit.max_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        ws_credit.add_chart(pie, "D1")

    ws_overdue = wb.create_sheet("Top Overdue Customers")
    ws_overdue.append(["Customer", "Credit Level", "Overdue Count", "Overdue Amount"])
    for cell in ws_overdue[1]:
        cell.font = header_font
        cell.fill = header_fill

    for c in stats.get("top_overdue_customers", []):
        ws_overdue.append([
            c["customer_name"], c["credit_level"],
            c["overdue_cnt"], c["overdue_total"]
        ])
        for cell in ws_overdue[ws_overdue.max_row]:
            cell.border = thin_border

    ws_trend = wb.create_sheet("Monthly Trend")
    ws_trend.append(["Month", "Completed", "Overdue"])
    for cell in ws_trend[1]:
        cell.font = header_font
        cell.fill = header_fill

    for t in stats.get("monthly_trend", []):
        ws_trend.append([t["month"], t["completed"], t["overdue"]])

    if stats.get("monthly_trend"):
        bar = BarChart()
        bar.title = "Monthly Payment Trend"
        bar.type = "col"
        bar.y_axis.title = "Count"
        data = Reference(ws_trend, min_col=2, max_col=3, min_row=1, max_row=ws_trend.max_row)
        cats = Reference(ws_trend, min_col=1, min_row=2, max_row=ws_trend.max_row)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws_trend.add_chart(bar, "E1")

    detail_headers = ["Contract No", "Customer", "Description",
                      "Planned Date", "Actual Date", "Amount",
                      "Overdue Days", "Status"]
    detail_col_widths = [18, 20, 16, 14, 14, 14, 14, 12]

    detail_configs = [
        ("Planned Payments", stats.get("planned_details", [])),
        ("Actual Receipts", stats.get("actual_details", [])),
        ("Overdue Unpaid", stats.get("overdue_details", [])),
    ]

    for sheet_title, details in detail_configs:
        ws_detail = wb.create_sheet(sheet_title)
        ws_detail.append(detail_headers)
        for cell in ws_detail[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for d in details:
            ws_detail.append([
                d.get("contract_no", ""),
                d.get("customer_name", ""),
                d.get("description", ""),
                d.get("planned_date", ""),
                d.get("actual_date", ""),
                d.get("amount", 0),
                d.get("overdue_days", 0),
                d.get("status", ""),
            ])
            for cell in ws_detail[ws_detail.max_row]:
                cell.border = thin_border

        for idx, width in enumerate(detail_col_widths, 1):
            ws_detail.column_dimensions[chr(64 + idx)].width = width

    customer_analysis = stats.get("customer_payment_analysis", [])
    if customer_analysis:
        ws_ca = wb.create_sheet("Customer Payment Analysis")
        ca_headers = ["Customer", "Credit Level", "Sales Manager",
                      "Month Planned", "Month Received", "Month Overdue",
                      "Payment Rate %", "Credit Change"]
        ws_ca.append(ca_headers)
        for cell in ws_ca[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for cp in customer_analysis:
            ws_ca.append([
                cp.get("customer_name", ""),
                cp.get("credit_level", ""),
                cp.get("sales_manager", ""),
                cp.get("month_planned", 0),
                cp.get("month_received", 0),
                cp.get("month_overdue", 0),
                cp.get("payment_rate", 0),
                cp.get("credit_change", "-"),
            ])
            for cell in ws_ca[ws_ca.max_row]:
                cell.border = thin_border

        ca_widths = [20, 12, 14, 14, 14, 14, 14, 20]
        for idx, width in enumerate(ca_widths, 1):
            ws_ca.column_dimensions[chr(64 + idx)].width = width

    manager_analysis = stats.get("manager_payment_analysis", [])
    if manager_analysis:
        ws_ma = wb.create_sheet("Manager Payment Analysis")
        ma_headers = ["Sales Manager", "Customer Count", "Frozen Count",
                      "Month Planned", "Month Received", "Month Overdue",
                      "Payment Rate %", "Collection Rate %"]
        ws_ma.append(ma_headers)
        for cell in ws_ma[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for ma in manager_analysis:
            ws_ma.append([
                ma.get("sales_manager", ""),
                ma.get("customer_count", 0),
                ma.get("frozen_count", 0),
                ma.get("month_planned", 0),
                ma.get("month_received", 0),
                ma.get("month_overdue", 0),
                ma.get("payment_rate", 0),
                ma.get("collection_rate", 0),
            ])
            for cell in ws_ma[ws_ma.max_row]:
                cell.border = thin_border

        ma_widths = [14, 14, 12, 14, 14, 14, 14, 14]
        for idx, width in enumerate(ma_widths, 1):
            ws_ma.column_dimensions[chr(64 + idx)].width = width

    for ws in wb.worksheets:
        if ws.title in ("Planned Payments", "Actual Receipts", "Overdue Unpaid",
                         "Customer Payment Analysis", "Manager Payment Analysis"):
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    col_letter = cell.column_letter
                    current = ws.column_dimensions[col_letter].width or 8
                    needed = min(len(str(cell.value)) + 5, 40)
                    if needed > current:
                        ws.column_dimensions[col_letter].width = needed

    wb.save(excel_path)
    logger.info(f"Excel report generated: {excel_path}")
    return excel_path


def _log_operation(op_type, contract_id, customer_id, details):
    execute_update(
        "INSERT INTO audit_logs (operation_type, contract_id, customer_id, operator, details, created_at) "
        "VALUES (?,?,?,?,?,datetime('now','localtime'))",
        [op_type, contract_id, customer_id, "system", details]
    )
