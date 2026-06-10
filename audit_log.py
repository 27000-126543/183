import csv
import json
import logging
from datetime import datetime
from models import execute_query, execute_update

logger = logging.getLogger(__name__)


def log_operation(operation_type, contract_id="", customer_id="", operator="system", details=""):
    if isinstance(details, dict):
        details = json.dumps(details, ensure_ascii=False)

    execute_update(
        "INSERT INTO audit_logs (operation_type, contract_id, customer_id, operator, details, created_at) "
        "VALUES (?,?,?,?,?,datetime('now','localtime'))",
        [operation_type, contract_id, customer_id, operator, details]
    )


def query_logs(contract_id=None, customer_id=None, operation_type=None,
               start_date=None, end_date=None, operator=None, limit=1000):
    sql = "SELECT * FROM audit_logs WHERE 1=1"
    params = []

    if contract_id:
        sql += " AND contract_id = ?"
        params.append(contract_id)
    if customer_id:
        sql += " AND customer_id = ?"
        params.append(customer_id)
    if operation_type:
        sql += " AND operation_type = ?"
        params.append(operation_type)
    if start_date:
        sql += " AND created_at >= ?"
        params.append(start_date)
    if end_date:
        sql += " AND created_at <= ?"
        params.append(end_date)
    if operator:
        sql += " AND operator = ?"
        params.append(operator)

    sql += " ORDER BY created_at DESC LIMIT ?"
    params.append(limit)

    return execute_query(sql, params, fetch_all=True)


def export_logs_csv(logs, output_path):
    if not logs:
        logger.warning("No logs to export")
        return None

    fieldnames = ["log_id", "operation_type", "contract_id", "customer_id",
                  "operator", "details", "created_at"]

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for log in logs:
            row = {k: log.get(k, "") for k in fieldnames}
            if isinstance(row["details"], dict):
                row["details"] = json.dumps(row["details"], ensure_ascii=False)
            writer.writerow(row)

    logger.info(f"Logs exported to CSV: {output_path}")
    return output_path


def export_logs_excel(logs, output_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        logger.error("openpyxl not installed, falling back to CSV")
        return export_logs_csv(logs, output_path.replace(".xlsx", ".csv"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["ID", "Operation Type", "Contract ID", "Customer ID",
               "Operator", "Details", "Created At"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for log in logs:
        row = [
            log.get("log_id", ""),
            log.get("operation_type", ""),
            log.get("contract_id", ""),
            log.get("customer_id", ""),
            log.get("operator", ""),
            json.dumps(log.get("details", {}), ensure_ascii=False) if isinstance(log.get("details"), dict) else str(log.get("details", "")),
            log.get("created_at", "")
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = thin_border

    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_letter = cell.column_letter
                current = ws.column_dimensions[col_letter].width or 8
                needed = min(len(str(cell.value)) + 3, 50)
                if needed > current:
                    ws.column_dimensions[col_letter].width = needed

    wb.save(output_path)
    logger.info(f"Logs exported to Excel: {output_path}")
    return output_path


def get_log_statistics(start_date=None, end_date=None):
    sql = "SELECT operation_type, COUNT(*) as cnt FROM audit_logs WHERE 1=1"
    params = []
    if start_date:
        sql += " AND created_at >= ?"
        params.append(start_date)
    if end_date:
        sql += " AND created_at <= ?"
        params.append(end_date)
    sql += " GROUP BY operation_type ORDER BY cnt DESC"

    return execute_query(sql, params, fetch_all=True)


def batch_export_logs(output_dir, contract_id=None, customer_id=None,
                      start_date=None, end_date=None, format_type="excel"):
    import os
    os.makedirs(output_dir, exist_ok=True)

    logs = query_logs(contract_id=contract_id, customer_id=customer_id,
                      start_date=start_date, end_date=end_date, limit=10000)

    if not logs:
        logger.info("No logs found for export criteria")
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"audit_logs_{timestamp}"

    if format_type == "excel":
        output_path = os.path.join(output_dir, f"{filename}.xlsx")
        return export_logs_excel(logs, output_path)
    else:
        output_path = os.path.join(output_dir, f"{filename}.csv")
        return export_logs_csv(logs, output_path)
